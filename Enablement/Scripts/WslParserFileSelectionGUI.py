# --    (c) WhereScape Inc 2020. WhereScape Inc permits you to copy this module solely for use with the RED software, and to modify this module            -- #
# --    for the purposes of using that modified module with the RED software, but does not permit copying or modification for any other purpose.           -- #
# --                                                                                                                                                       -- #
#=====================================================================================================
# Module Name      :    WslParserFileSelectionGUI
# DBMS Name        :    Generic for all databases
# Description      :    Generic python functions module used by Browse_File_Parser.py 
#                       The Module contains functions to generate GUI to select file profile options and file preview
# Author           :    Wherescape Inc
#======================================================================================================
# Notes / History
# 1.0.0   2022-02-16   First Version
#======================================================================================================

import os
import tkinter as tk
from tkinter import *
import tkinter.font as tkFont
from tkinter import filedialog
import tkinter.ttk as ttk
import os
from itertools import islice
from tkinter import scrolledtext
import pandas as pd
from idlelib.tooltip import Hovertip
from tkinter import messagebox                              
from openpyxl import load_workbook
import fastavro
import pyorc

parserType = ('JSON','DELIMITED','AVRO','PARQUET','ORC','XLSX', 'XML')
encodingTypes = ('UTF-8', 'ANSI', 'UTF-16')
def on_closing():
    if messagebox.askyesno('Quit', 'Do you want to quit?'):
        top.destroy()
        boot.destroy()                

def createWindow():
    global boot
    boot = tk.Tk()
    boot.withdraw()
    global varSheetName
    global varEncoding
    varEncoding = tk.StringVar()
    varSheetName = tk.StringVar()
    return boot

def createLabel(top,textLabel,x1,y1,width1,height1,tooltipText=''):
    addLabel = tk.Label(top)
    Hovertip(addLabel,tooltipText)
    ft = tkFont.Font(family='Tahoma', size=10)
    addLabel['font'] = ft
    addLabel['fg'] = '#333333'
    addLabel['justify'] = 'left'
    addLabel['text'] = textLabel
    addLabel.place(x=x1, y=y1, width=width1, height=height1)
    return addLabel

def createFileNameLabel(top,textLabel):
    top.title(f'Options to Parse File: {textLabel}')

def createCheckBoxForSaveOptions(top,textLabel,x1,y1,width1,height1):
    global varAddCheckBox
    varAddCheckBox = tk.IntVar(value='0')
    addCheckBox = tk.Checkbutton(top, variable=varAddCheckBox)						  
    ft = tkFont.Font(family='Tahoma', size=10)
    addCheckBox['font'] = ft
    addCheckBox['fg'] = '#333333'
    addCheckBox['justify'] = 'center'
    addCheckBox['text'] = textLabel
    addCheckBox.place(x=x1, y=y1, width=width1, height=height1)
    addCheckBox['offvalue'] = '0'
    addCheckBox['onvalue'] = '1'
    return varAddCheckBox,addCheckBox


def createCheckBox(top,textLabel,x1,y1,width1,height1,fileName='',i='',defaultVal=0,state='normal',tooltipText=''):
    global varAddCheckBox
    varAddCheckBox = tk.IntVar(value='0')
    addCheckBox = tk.Checkbutton(top, variable=varAddCheckBox, state=state, command=lambda : checkBoxClick(fileName,i,varAddCheckBox))
    Hovertip(addCheckBox,tooltipText)
    varAddCheckBox.set(defaultVal)
    ft = tkFont.Font(family='Tahoma', size=10)
    addCheckBox['font'] = ft
    addCheckBox['fg'] = '#333333'
    addCheckBox['justify'] = 'center'
    addCheckBox['text'] = textLabel
    addCheckBox.place(x=x1, y=y1, width=width1, height=height1)
    addCheckBox['offvalue'] = '0'
    addCheckBox['onvalue'] = '1'
    return varAddCheckBox,addCheckBox

def createTextBox(top,x1,y1,width1,height1,varDefault,state, tooltipText=''):
    global varAddTextBox
    varAddTextBox = tk.StringVar()
    addTextBox = tk.Entry(top, textvariable=varAddTextBox, state=state)
    Hovertip(addTextBox,tooltipText)
    addTextBox.insert(0, varDefault)
    addTextBox['borderwidth'] = '1px'
    ft = tkFont.Font(family='Tahoma', size=10)
    addTextBox['font'] = ft
    addTextBox['fg'] = '#333333'
    addTextBox['justify'] = 'center'
    addTextBox.place(x=x1, y=y1, width=width1, height=height1)
    addTextBox.bind('<FocusIn>', callback)                                          
    return varAddTextBox,addTextBox
def callback(event):
    # select text
    event.widget.select_range(0, 'end')
    # move cursor to the end
    event.widget.icursor('end')
def createButton(top,textLabel,x1,y1,width1,height1,i,iconPath):
    addButton = tk.Button(top, relief=RIDGE, command=lambda : 
                          objectAllOptions(i,iconPath))
    addButton['bg'] = '#f0f0f0'
    addButton['bd'] = '1'
    ft = tkFont.Font(family='Tahoma', size=10)
    addButton['font'] = ft
    addButton['fg'] = 'black'
    addButton['justify'] = 'center'
    addButton['text'] = textLabel
    addButton.place(x=x1, y=y1, width=width1, height=height1)

def clearAllObjects():
      for obj in getOnlyObject:
          try:
           obj.destroy()
          except Exception:
           pass
      for widget in getOnlyLabel:
          try:
            widget.destroy()
          except Exception:
           pass    

def getFileOptions(fileName,i,varAddComboBox,headerlineStatus = 'disabled',iconPath=''):
    if (varAddComboBox == ''):
        if "."  in fileName:
            fileExt = fileName.split('.')[-1]
            if fileExt in ['txt', 'csv', 'dat']:
                fileExt = 'delimited'
            elif fileExt in ['xls', 'xlsx']:
                fileExt = 'xlsx'
            elif fileExt in ['orc']:
                fileExt = 'orc'
            elif fileExt in ['avro']:
                fileExt = 'avro'
            elif fileExt in ['json']:
                fileExt = 'json'
            elif fileExt in ['parquet']:
                fileExt = 'parquet'
            elif fileExt in ['xml']:
                fileExt = 'xml'
            else:
                fileExt = 'delimited'
        else:
            fileExt = 'delimited'
    else:
        fileExt=varAddComboBox.get().lower().strip()

    for winObject in allObjectDictionary[fileExt]:
        dictAllObjectValues.update({'File Name': fileName})
        if winObject['objType'] == 'bn':
            createButton(top,winObject['Label'],winObject['x'],winObject['y'],winObject['h'],winObject['w'],i,iconPath='')
        elif winObject['objType'] == 'tb':
            tempLabelObject=createLabel(top,winObject['Label'],winObject['x'],winObject['y'],winObject['h'],winObject['w'],winObject['tooltip'])

            tempReturnObject = createTextBox(
                top,winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],winObject['default'], winObject['state'],winObject['tooltip'])
            dictAllObjectValues.update({winObject['Label']: tempReturnObject[0]})
            getOnlyObject.append(tempReturnObject[1])
            getOnlyLabel.append(tempLabelObject)
        
        elif winObject['objType'] == 'cb':
            tempLabelObject=createLabel(top,winObject['Label'],winObject['x'],winObject['y'],winObject['h'],winObject['w'],winObject['tooltip'])
            if winObject['Label'] == "Header Record":
                winObject['defaultVal'] = varAddCheckBox.get()
            tempReturnObject = createCheckBox(
                top,winObject['default'],winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],fileName,i,winObject['defaultVal'],winObject['state'], winObject['tooltip'])
            dictAllObjectValues.update({winObject['Label']: tempReturnObject[0]})
            getOnlyObject.append(tempReturnObject[1])
            getOnlyLabel.append(tempLabelObject)

        elif winObject['objType'] == 'db':
            tempLabelObject=createLabel(top,winObject['Label'],winObject['x'],winObject['y'],winObject['h'],winObject['w'],winObject['tooltip'])
            
            if winObject['Label'] == 'SheetName':
                sheetNameComboBox(top,winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],winObject['default'],winObject['defaultSelected'],fileName,i, winObject['state'], winObject['tooltip'])
            elif winObject['Label'] == 'Encoding Type':
                encodingComboBox(top,winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],winObject['default'],winObject['defaultSelected'],fileName,i, winObject['state'], winObject['tooltip'])
            else:
                tempReturnObject = createComboBox(top,winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],winObject['default'],winObject['defaultSelected'],fileName,i,winObject['state'], winObject['tooltip'],iconPath)
            dictAllObjectValues.update({winObject['Label']: tempReturnObject[0]})
            if(winObject['Label'] == 'SheetName'):
                getOnlyObject.append(tempReturnObject[1])
                getOnlyLabel.append(tempLabelObject)


def selectComboOptions(fileName,i,varAddComboBox,iconPath=''):
    clearAllObjects()    
    getFileOptions(fileName,i,varAddComboBox,iconPath)

def checkBoxClick(fileName,i,varAddCheckBox):
    getExtension=fileName.split(".")[-1]
    if getExtension in ['xls','xlsx']:                                                                              
     if varAddCheckBox.get() == 1:
        headerlineStatus = 'normal'
     else:
        headerlineStatus = 'disabled'
     for winObject in allObjectDictionary[fileExt]:
        if winObject['objType'] == 'tb':
            tempLabelObject=createLabel(top,winObject['Label'],winObject['x'],winObject['y'],winObject['h'],winObject['w'],winObject['tooltip'])
            if winObject['Label'] == "Header Lines":
                winObject['state'] = headerlineStatus
                tempReturnObject = createTextBox(
                top,winObject['x1'],winObject['y1'],winObject['h1'],winObject['w1'],winObject['default'], winObject['state'],winObject['tooltip'])
                dictAllObjectValues.update({winObject['Label']: tempReturnObject[0]})
                getOnlyObject.append(tempReturnObject[1])
                getOnlyLabel.append(tempLabelObject)

def selectSheetNameCombo(e,fileName):
    createFilePreview(top,20,220,460,220,fileName,sheetName=varSheetName.get())
    
def createComboBox(top,x1,y1,width1,height1,varDefaultAll,varDefault,fileName,i,state,tooltipText='',iconPath=''):
    global varAddComboBox
    varAddComboBox = tk.StringVar()
    addComboBox = ttk.Combobox(top, textvariable=varAddComboBox, state=state)
    Hovertip(addComboBox,tooltipText)
    addComboBox['values'] = (varDefaultAll)
    ft = tkFont.Font(family='Tahoma', size=8)
    addComboBox['font'] = ft
    addComboBox['justify'] = 'center'
    addComboBox.current(varDefault)
    addComboBox.place(x=x1, y=y1, width=width1, height=height1)
    addComboBox.bind('<<ComboboxSelected>>',  lambda e:selectComboOptions(fileName,i,varAddComboBox,iconPath))
    return varAddComboBox,addComboBox


def sheetNameComboBox(top,x1,y1,width1,height1,varDefaultAll,varDefault,fileName,i,state,tooltipText=''):
    sheetNameCombo = ttk.Combobox(top, textvariable=varSheetName, state=state)
    Hovertip(sheetNameCombo,tooltipText)
    sheetNameCombo['values'] = (varDefaultAll)
    ft = tkFont.Font(family='Tahoma', size=8)
    sheetNameCombo['font'] = ft
    sheetNameCombo['justify'] = 'center'
    sheetNameCombo.current(varDefault)
    sheetNameCombo.place(x=x1, y=y1, width=width1, height=height1)
    sheetNameCombo.bind('<<ComboboxSelected>>',  lambda e:selectSheetNameCombo(e, fileName))

def encodingComboBox(top,x1,y1,width1,height1,varDefaultAll,varDefault,fileName,i,state,tooltipText=''):
    encodingCombo = ttk.Combobox(top, textvariable=varEncoding, state=state)
    Hovertip(encodingCombo,tooltipText)
    encodingCombo['values'] = (varDefaultAll)
    ft = tkFont.Font(family='Tahoma', size=8)
    encodingCombo['font'] = ft
    encodingCombo['justify'] = 'center'
    encodingCombo.current(varDefault)
    encodingCombo.place(x=x1, y=y1, width=width1, height=height1)

def createFilePreview(top,x1,y1,width1,height1,fileName,sheetName=''):
    framePreview = Frame(top)
    framePreview.place(x=10,y=250, width=480, height=220)
    h=Scrollbar(framePreview, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')
    global addFilePreview
    addFilePreview = scrolledtext.ScrolledText(framePreview, width=140, height=120, wrap = NONE, xscrollcommand=h.set)
    addFilePreview.configure(font=("Tahoma", 10))
    addFilePreview.pack(side=BOTTOM)
  
    try:
        if fileName.split('.')[-1] in ['txt', 'csv', 'dat', 'json','xml']:
            with open(fileName, 'r') as input_file:
                lines_cache = islice(input_file, 100)
                for current_line in lines_cache:
                    addFilePreview.insert(tk.END, str(current_line))

        elif fileName.split('.')[-1] in ['orc']:
            with open(fileName, 'rb') as fp:
                reader = pyorc.Reader(fp)
                records = [r for r in reader]
                df = pd.DataFrame.from_records(records)
                addFilePreview.insert(END, df)

        elif fileName.split('.')[-1] in ['avro']:
            with open(fileName, 'rb') as fp:
                reader = fastavro.reader(fp)
                records = [r for r in reader]
                df = pd.DataFrame.from_records(records)
                addFilePreview.insert(END, df)

        elif fileName.split('.')[-1] in ['parquet']:
            df = pd.read_parquet(fileName,engine='auto')
            addFilePreview.insert(END, df.head())


        elif fileName.split('.')[-1] in ['xlsx']:
            wb = load_workbook(fileName, read_only=True, keep_links=False)
            sheet_names = wb.sheetnames
            if sheetName == '':
                sheetName = sheet_names[0]
            else:
                sheetName = sheetName
            df = pd.read_excel(fileName, sheet_name=sheetName)
            addFilePreview.insert(END, df.head())
        else:
            addFilePreview.insert(tk.END, "No Preview Available")
    except Exception as e:
        addFilePreview.insert(tk.END, "No Preview Available")
        pass

    h.config(command=addFilePreview.xview)
    addFilePreview.configure(state='disabled')

def createParseWindow(boot, fileName,iconPath=''):
    global top
    top = Toplevel(boot)
    top.grab_set()
    top.geometry('750x250')
    top.protocol("WM_DELETE_WINDOW", on_closing)                                           
    # setting window size
    width = 500
    height = 500
    screenwidth = top.winfo_screenwidth()
    screenheight = top.winfo_screenheight()
    alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width)
                                / 2, (screenheight - height) / 2)
    top.geometry(alignstr)
    top.resizable(width=False, height=False)

    top.iconbitmap(iconPath)
    return top


def selectFilesFromWindows():
    # global selectedFilesList
    selectedFilesList = list(filedialog.askopenfilenames(title='Choose a file',filetypes=[('All Files', '*.*')]))
    return selectedFilesList

def objectAllOptions(i,iconPath=''):
    fileParser=dictAllObjectValues['Parser'].get().lower().strip()
    if fileParser == 'delimited':
        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for allFilename in selectedFilesList[i:]:
                parserDictionary = {
                    'fieldDelimiter': dictAllObjectValues['Field Delimiter'].get() or "|",                 
                    'recordDelimiter': dictAllObjectValues['Record Delimiter'].get() or "\n", 
                    'enclosedBy': dictAllObjectValues['Field Enclosure Delimiter'].get() or '"',             
                    'headerLine': dictAllObjectValues['Header Lines'].get() or "0",               
                    'skipLine': dictAllObjectValues['Skip Lines'].get() or "0",               
                    'checkBoxHeader': dictAllObjectValues['Header Record'].get(),                
                    'rowLimit': dictAllObjectValues['Row Limit For Profiling'].get() or "100",                  
                    'parser': dictAllObjectValues['Parser'].get(),                
                    'fileName': allFilename,               
                    'encoding': varEncoding.get(),
                    }
                fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary = {
                'fieldDelimiter': dictAllObjectValues['Field Delimiter'].get() or "|",           
                'recordDelimiter': dictAllObjectValues['Record Delimiter'].get() or "\n",             
                'enclosedBy': dictAllObjectValues['Field Enclosure Delimiter'].get() or '"',             
                'headerLine': dictAllObjectValues['Header Lines'].get() or "0",             
                'skipLine': dictAllObjectValues['Skip Lines'].get() or "0",              
                'checkBoxHeader': dictAllObjectValues['Header Record'].get(),            
                'rowLimit': dictAllObjectValues['Row Limit For Profiling'].get() or "100",              
                'parser': dictAllObjectValues['Parser'].get(),                
                'fileName': dictAllObjectValues['File Name'],            
                'encoding': varEncoding.get()
                }
            fileParseInfo.append(parserDictionary)
            top.destroy()

    elif fileParser == 'xlsx':
        
        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for index, allFilename in enumerate(selectedFilesList[i:]):
                if index == 0:
                    parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(),'SheetName': varSheetName.get(),'headerLine': dictAllObjectValues['Header Lines'].get() or "0",'checkBoxHeader': dictAllObjectValues['Header Record'].get(),'rowLimit': dictAllObjectValues['Row Limit For Profiling'].get()or "100",'encoding': varEncoding.get()}
                    fileParseInfo.append(parserDictionary)
                else:
                    parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(),'SheetName': "DEF_0",'headerLine': dictAllObjectValues['Header Lines'].get() or "0",'checkBoxHeader': dictAllObjectValues['Header Record'].get(),'rowLimit': dictAllObjectValues['Row Limit For Profiling'].get()or "100", 'encoding': varEncoding.get()}
                    fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary = {'fileName': dictAllObjectValues['File Name'],'parser': dictAllObjectValues['Parser'].get(),'SheetName': varSheetName.get(),'headerLine': dictAllObjectValues['Header Lines'].get()or "0",'checkBoxHeader': dictAllObjectValues['Header Record'].get(),'rowLimit': dictAllObjectValues['Row Limit For Profiling'].get()or "100", 'encoding': varEncoding.get()}
            fileParseInfo.append(parserDictionary)
            top.destroy()

    elif fileParser in ['orc','parquet']:

        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for allFilename in selectedFilesList[i:]:
                parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(), 'Profiling factor':dictAllObjectValues['Profiling factor'].get(), 'encoding': varEncoding.get()}
                fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary =  {'fileName': dictAllObjectValues['File Name'],'parser': dictAllObjectValues['Parser'].get(), 'Profiling factor':dictAllObjectValues['Profiling factor'].get()or "1.5", 'encoding': varEncoding.get()}
            fileParseInfo.append(parserDictionary)
            top.destroy()

    elif fileParser == 'json':

        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for allFilename in selectedFilesList[i:]:
                parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(), 'Depth' :dictAllObjectValues['Depth'].get()or "1.5", 'encoding': varEncoding.get()}
                fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary =  {'fileName': dictAllObjectValues['File Name'],'parser': dictAllObjectValues['Parser'].get(), 'Depth' :dictAllObjectValues['Depth'].get()or "1.5", 'encoding': varEncoding.get()}
            fileParseInfo.append(parserDictionary)
            top.destroy()
    
    elif fileParser == 'xml':

        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for allFilename in selectedFilesList[i:]:
                parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(),'Depth' :dictAllObjectValues['Depth'].get()or "1.5",'encoding': varEncoding.get()}
                fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary =  {'fileName': dictAllObjectValues['File Name'],'parser': dictAllObjectValues['Parser'].get(), 'Depth' :dictAllObjectValues['Depth'].get()or "1.5", 'encoding': varEncoding.get()}
            fileParseInfo.append(parserDictionary)
            top.destroy()
            
    else:

        if varcheckBoxSameOptions[0].get() == 1:
            i = i - 1
            for allFilename in selectedFilesList[i:]:
                parserDictionary = {'fileName': allFilename, 'parser': dictAllObjectValues['Parser'].get(),'encoding': varEncoding.get()}
                fileParseInfo.append(parserDictionary)
            boot.destroy()
            return fileParseInfo
        else:
            parserDictionary =  {'fileName': dictAllObjectValues['File Name'],'parser': dictAllObjectValues['Parser'].get(),'encoding': varEncoding.get()}
            fileParseInfo.append(parserDictionary)
            top.destroy()
    if i < len(selectedFilesList):
        getParseOptions(boot, selectedFilesList[i], i,iconPath)
    else:
        boot.destroy()
    return fileParseInfo


def getAllObjects():
    global allObjectDictionary


    allObjectDictionary = {
        'delimited': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'normal'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'normal'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'normal'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  1 ,'tooltip':'Parser Type', 'state':'readonly'},{
                    'objType': 'tb', 'Label': 'Header Lines', 'x': 280, 'y': 120, 'h': 100, 'w': 25, 'x1': 400, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '0', 'tooltip': 'A number which informs the parser of the line number on which the header resides.', 'state': 'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'normal'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'normal'},{
                  'objType': 'tb','Label': 'Row Limit For Profiling','x': 28,'y': 120,'h': 150,'w': 25,'x1': 200,'y1': 120,'h1': 70,'w1': 25,'default': '100','tooltip':'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state':'normal'},{
                  'objType': 'bn','Label': 'OK','x': 270,'y': 225,'h': 200,'w': 25,}, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'normal'}, {
                  'objType': 'tb','Label': 'Profiling factor','x': 285,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1.5,'tooltip':'Multiplier applied to profiled data lengths to allow for some data length variation', 'state':'disabled'}
                ),
        'xlsx': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.\n A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  5 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'normal'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'normal'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Profiling factor','x': 285,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1.5,'tooltip':'Multiplier applied to profiled data lengths to allow for some data length variation', 'state':'disabled'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'normal'}
                ),

        'json': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  0 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'disabled'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Depth','x': 274,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1,'tooltip':'Number of levels in which JSON will be parsed.\n Note: Time taken for parsing will increase for larger value.', 'state':'normal'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'normal'}
                ),
        'avro': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  2 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'disabled'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Profiling factor','x': 285,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1.5,'tooltip':'Multiplier applied to profiled data lengths to allow for some data length variation', 'state':'disabled'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'disabled'}
                ),
       'parquet': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  3 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'disabled'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Profiling factor','x': 285,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1.5,'tooltip':'Multiplier applied to profiled data lengths to allow for some data length variation', 'state':'normal'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'normal'}
                ),    
        'orc': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  4 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'disabled'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Profiling factor','x': 285,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 1.5,'tooltip':'Multiplier applied to profiled data lengths to allow for some data length variation', 'state':'normal'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'disabled'}
                ),
        'xml': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  6 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                  'objType': 'tb','Label': 'Row Limit For Profiling','x': 28,'y': 120,'h': 150,'w': 25,'x1': 200,'y1': 120,'h1': 70,'w1': 25,'default': '100','tooltip':'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state':'disabled'},{
                  'objType': 'bn','Label': 'OK','x': 270,'y': 225,'h': 200,'w': 25,}, {
                'objType': 'db','Label': 'SheetName','x': 20,'y': 150,'h': 100,'w': 25,'x1': 200,'y1': 150,'h1': 70,'w1': 25,'default': (excelSheets),'defaultSelected' :  0,'tooltip':'Sheetname from xlsx file', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Depth','x': 274,'y': 150,'h': 100,'w': 25,'x1': 400,'y1': 150,'h1': 70,'w1': 25,'default': 5,'tooltip':'Number of levels in which XML will be parsed.', 'state':'normal'},{
                'objType': 'db','Label': 'Encoding Type','x': 29,'y': 180,'h': 100,'w': 25,'x1': 200,'y1': 180,'h1': 70,'w1': 25,'default': (encodingTypes),'defaultSelected' :  0,'tooltip':'Encoding type for file', 'state':'normal'}
                ),
        'other': (
                { 'objType': 'tb','Label': 'Field Delimiter','x': 29,'y': 30,'h': 100,'w': 25,'x1': 200,'y1': 30,'h1': 70,'w1': 25,'default': '|','tooltip':'Character that separates the fields within each record of the source file.The field delimiter identifies end of each field. \n Common Field Delimiters are comma,tab,colon.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Record Delimiter','x': 35,'y': 60,'h': 100,'w': 25,'x1': 200,'y1': 60,'h1': 70,'w1': 25,'default': '\\n','tooltip':'String to identify how each line/record in source file is ended/terminated/delineated.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Field Enclosure Delimiter','x': 33,'y': 90,'h': 150,'w': 25,'x1': 200,'y1': 90,'h1': 70,'w1': 25,'default': '"','tooltip':'Character that delimits BOTH start and end of field value i.e. encapsulates value.A double quote is common enclosure delimiter', 'state':'disabled'},{
                  'objType': 'db','Label': 'Parser','x': 273,'y': 30,'h': 80,'w': 25,'x1': 400,'y1': 30,'h1': 90,'w1': 25,'default': parserType,'defaultSelected' :  1 ,'tooltip':'Parser Type', 'state':'readonly'},{
                  'objType': 'tb','Label': 'Header Lines','x': 280,'y': 120,'h': 100,'w': 25,'x1': 400,'y1': 120,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which informs the parser of the line number on which the header resides.', 'state':'disabled'},{
                  'objType': 'tb','Label': 'Skip Lines','x': 273,'y': 60,'h': 100,'w': 25,'x1': 400,'y1': 60,'h1': 70,'w1': 25,'default': '0','tooltip':'A number which specifies the parser the line number from which data profiling should start.\n First line of data set starts from 0', 'state':'disabled'},  {
                  'objType': 'cb','Label': 'Header Record','x': 287,'y': 90,'h': 100,'w': 25,'x1': 400,'y1': 90,'h1': 70,'w1': 25,'default': '','tooltip':'Indicates whether the source file contains a heading/label for each field, \n which is not regarded as data so it should not be loaded','defaultVal':0 , 'state':'disabled'},{
                    'objType': 'tb', 'Label': 'Row Limit For Profiling', 'x': 28, 'y': 120, 'h': 150, 'w': 25, 'x1': 200, 'y1': 120, 'h1': 70, 'w1': 25, 'default': '100', 'tooltip': 'Number of records to scan for Data Profiling. \n NOTE: This is for analysis of data for datatypes identification.', 'state': 'disabled'}, {
                    'objType': 'bn', 'Label': 'OK', 'x': 270, 'y': 225, 'h': 200, 'w': 25, }, {
                    'objType': 'db', 'Label': 'SheetName', 'x': 20, 'y': 150, 'h': 100, 'w': 25, 'x1': 200, 'y1': 150, 'h1': 70, 'w1': 25, 'default': (excelSheets), 'defaultSelected':  0, 'tooltip': 'Sheetname from xlsx file', 'state': 'disabled'}, {
                    'objType': 'db', 'Label': 'Encoding Type', 'x': 29, 'y': 180, 'h': 100, 'w': 25, 'x1': 200, 'y1': 180, 'h1': 70, 'w1': 25, 'default': (encodingTypes), 'defaultSelected':  0, 'tooltip': 'Encoding type for file', 'state': 'normal'}
                )}


def getParseOptions(boot, fileName, i,iconPath=''):
    global dictAllObjectValues, parserDictionary, varcheckBoxSameOptions,getOnlyObject,getOnlyLabel,fileExt,excelSheets
    excelSheets=['None']
    dictAllObjectValues = {}
    getOnlyObject=[]
    getOnlyLabel=[]
    top = createParseWindow(boot, selectedFilesList[i],iconPath)

    createFileNameLabel(top, selectedFilesList[i].split("/")[-1])
    varcheckBoxSameOptions = createCheckBoxForSaveOptions(top, 'Save Options For All Files', 20, 225, 200, 25)
    fileExt = fileName.split('.')[-1]
    createFilePreview(top,20,220,460,230,selectedFilesList[i])
    if fileExt in ['txt', 'csv', 'dat']:
        fileExt = 'delimited'
    if fileExt in ['xls','xlsx']:
            xl = pd.ExcelFile(fileName)
            excelSheets=xl.sheet_names
    i = i + 1
    getAllObjects()
    getFileOptions(fileName,i,'','disabled',iconPath)

def parseSelectedFiles(boot,fileList,iconPath=''):
    global selectedFilesList
    if fileList == None:
        selectedFilesList = selectFilesFromWindows()
    else:
        fileList = [file.replace(os.sep, '/') for file in fileList]
        selectedFilesList = fileList
    getParseOptions(boot, selectedFilesList[0], 0,iconPath)

global fileParseInfo
global i
fileParseInfo = []

def getSelectedFiles(iconPath,fileList=None):
    try:
        boot = createWindow()
        parseSelectedFiles(boot,fileList,iconPath)
        boot.mainloop()
    except Exception as e:
        fileParseInfo.append("Error")
        pass

    return fileParseInfo